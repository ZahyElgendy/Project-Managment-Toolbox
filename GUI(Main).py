import PERT
import GANTT_CHART as gnt
import Network_Diagram as nd
import Cost


from pathlib import Path
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage, filedialog, Label, ttk



OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(r"GUI_Pictures\frame0")

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)


window = Tk()

window.geometry("720x576")
window.configure(bg = "#FFFFFF")




def main():


    def load_excel(path):  # Excel Widget Code inspired by Youtube Channel "Code first with Hala" https://www.youtube.com/watch?v=cTqP7xLk7j4&t=371s
        import sys
        import openpyxl
        from PyQt5.QtWidgets import (
            QApplication,
            QWidget,
            QVBoxLayout,
            QTableWidget,
            QComboBox,
            QPushButton,
            QTableWidgetItem
        )
        

        class ExcelViewer(QWidget):   #Defines new class that inherts from Qwidget
            def __init__(self, excel_path): #initialization of attributes of window
                super().__init__()
                self.setWindowTitle("Excel Viewer")  #Sets the title of application window

                #Stores and loades the excel file path
                self.excel_path = excel_path
                self.workbook = openpyxl.load_workbook(excel_path)

                #Main Layout
                self.layout = QVBoxLayout()   #Lays elements in one vertical coloumn 
                self.setLayout(self.layout)

                #Creat table widget
                self.table_widget = QTableWidget()
                #adds the window widget to the layout
                self.layout.addWidget(self.table_widget)

                #Create combobox to select sheets
                self.sheet_combo_box = QComboBox()
                self.sheet_combo_box.addItems(self.workbook.sheetnames) #Add sheet names to combobox
                self.layout.addWidget(self.sheet_combo_box)

                #Adds a button named "Load Sheet"
                self.load_button = QPushButton("Load Sheet")
                #Connects the button "Load Sheet" with the load_data function in class
                self.load_button.clicked.connect(self.load_data)
                #Adds the button widget to layout
                self.layout.addWidget(self.load_button)

                self.load_data()  #By default, load the first sheet

            def load_data(self):
                sheet_name = self.sheet_combo_box.currentText()
                #To open the current sheet selected
                sheet = self.workbook[sheet_name]

                #To clear any previous data
                self.table_widget.clearContents()
                self.table_widget.setRowCount(0)

                #Sets the widget table rows and coloumns according to maximum rows and coloumns in sheet
                self.table_widget.setRowCount(sheet.max_row)
                self.table_widget.setColumnCount(sheet.max_column)

                #To convert all values of the sheets into a list of tuples
                list_values = list(sheet.values)
                
                #This will set the first row of the excel file as header
                self.table_widget.setHorizontalHeaderLabels(list_values[0])

                #Fill the table cells with data from 2nd row and onwards
                row_index = 0

                #To convert all values of the sheets into a list
                for value_tuple in list_values[1:]:
                    col_index = 0
                    for value in value_tuple:
                        self.table_widget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                        col_index += 1
                    row_index += 1


        #To execute the application
        if __name__ == "__main__":
            app = QApplication(sys.argv) #Define Qapplication iside variable named app, sys.argv is to recieve the arguments passed into it
            window = ExcelViewer(path)   #Pass the excel file to be loaded in the window
            window.showMaximized()       # launches application and maximizes to full screen
            app.exec_()                  #Execute the application



    canvas = Canvas(
        window,         #Initializes Parent Window
        bg = "#FFFFFF", #Sets background color to white (Code: #FFFFFF)
        height = 576,   #Parent Window Canvas Height
        width = 720,    #Parent Window Canvas Width
        bd = 0,         # Border width = 0 i.e., no border
        highlightthickness = 0,  #no highlight
        relief = "ridge" #border effect style
    )

    canvas.place(x = 0, y = 0) #This is to place the Canvas at the top left corner of the window
    
    #Placing image in canvas
    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1_1.png"))  #retrieves the image from the file
    image_1 = canvas.create_image( #Position of image 1
        360.0,
        288.0,
        image=image_image_1
    )

    image_image_3 = PhotoImage(
        file=relative_to_assets("image_3_2.png"))
    image_3 = canvas.create_image(
        320.0,
        88.0,
        image=image_image_3
    )


#Creating Buttons

#Browse Button
    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1_1.png")) #Retrieves button image
    button_1 = Button( 
        image=button_image_1, #Sets buttom image
        borderwidth=0,        #No button width
        highlightthickness=0, #No Highlight
        command=lambda: File_dialog(),  #Connect button with file_dialog function 
        relief="flat"         #Flat button style for aesthetic beauty
    )
    button_1.place(
        x=520.0,
        y=315.0,
        width=141.0,
        height=40.0
    )

#Similar for first button, but with different image and command
#Start Button
    button_image_2 = PhotoImage(
        file=relative_to_assets("button_2_1.png"))
    button_2 = Button(
        image=button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: [PERT.PERT(filename), screen_2()],
        relief="flat"
    )
    button_2.place(
        x=285.0,
        y=390.0,
        width=141.0,
        height=40.0
    )

#To Create Entry label for browse
    entry_image_1 = PhotoImage(
        file=relative_to_assets("entry_1_1.png"))  #Retrieve entry image from file
    entry_bg_1 = canvas.create_image(  #Position of entry label 
        296.5,
        337.0,
        image=entry_image_1
    )
    entry_1 = Label(
        bd=0,
        bg="#FFFFFF",  #Background White color
        fg="#000716",  #Forefround dark blue color
        highlightthickness=0,  # highlight thickness 0
        text="No File Selected" #default text 
    )
    entry_1.place(
        x=110.0,
        y=321.0,
        width=375.0,
        height=30.0
    )

    def File_dialog():
        global filename
        """This Function will open the file explorer window and assign the chosen file path to label_file"""
        filename = filedialog.askopenfilename(initialdir="/",
                                            title="Select A File",
                                            filetype=(("xlsx files", "*.xlsx"),("All Files", "*.*")))
        print(f"Loaded file path is : {filename}")
        entry_1["text"] = filename

        return None



#Similar to previous, the next screens' imnages and buttons  are created in a similar fashion
    def screen_2():

        canvas = Canvas(
            window,
            bg = "#FFFFFF",
            height = 576,
            width = 720,
            bd = 0,
            highlightthickness = 0,
            relief = "ridge"
        )

        canvas.place(x = 0, y = 0)
        image_image_1 = PhotoImage(
            file=relative_to_assets("image_1_2.png"))
        image_1 = canvas.create_image(
            360.0,
            288.0,
            image=image_image_1
        )

        image_image_2 = PhotoImage(
            file=relative_to_assets("image_2_2.png"))
        image_2 = canvas.create_image(
            347.0,
            312.0,
            image=image_image_2
        )

        image_image_3 = PhotoImage(
            file=relative_to_assets("image_3_2.png"))
        image_3 = canvas.create_image(
            320.0,
            88.0,
            image=image_image_3
        )

        button_image_1 = PhotoImage(
            file=relative_to_assets("button_1_2.png"))
        button_1 = Button(
            image=button_image_1,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: timeScreen(),
            relief="flat"
        )
        button_1.place(
            x=130.0,
            y=369.0,
            width=141.0,
            height=45.0
        )

        button_image_2 = PhotoImage(
            file=relative_to_assets("button_2_2.png"))
        button_2 = Button(
            image=button_image_2,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: [Cost.cost(r"PERT.xlsx"), budgetScreen()],
            relief="flat"
        )
        button_2.place(
            x=417.0,
            y=369.0,
            width=141.0,
            height=45.0
        )

        button_image_3 = PhotoImage(
            file=relative_to_assets("button_3_2.png"))
        button_3 = Button(
            image=button_image_3,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: main(),
            relief="flat"
        )
        button_3.place(
            x=25.0,
            y=497.0,
            width=90.0,
            height=73.0
        )






        def timeScreen():
            canvas = Canvas(
                window,
                bg = "#FFFFFF",
                height = 576,
                width = 720,
                bd = 0,
                highlightthickness = 0,
                relief = "ridge"
            )

            canvas.place(x = 0, y = 0)
            image_image_1 = PhotoImage(
                file=relative_to_assets("image_1_3.png"))
            image_1 = canvas.create_image(
                360.0,
                288.0,
                image=image_image_1
            )

            image_image_2 = PhotoImage(
                file=relative_to_assets("image_2_3.png"))
            image_2 = canvas.create_image(
                360.0,
                238.0,
                image=image_image_2
            )

            image_image_3 = PhotoImage(
                file=relative_to_assets("image_3_3.png"))
            image_3 = canvas.create_image(
                433.0,
                65.0,
                image=image_image_3
            )

            image_image_4 = PhotoImage(
                file=relative_to_assets("image_4_3.png"))
            image_4 = canvas.create_image(
                320.0,
                88.0,
                image=image_image_4
            )

        #GANTT button
            button_image_1 = PhotoImage(
                file=relative_to_assets("button_1_3.png"))
            button_1 = Button(
                image=button_image_1,
                borderwidth=0,
                highlightthickness=0,
                command=lambda: gnt.Gant("PERT.xlsx"),
                relief="flat"
            )
            button_1.place(
                x=286.0,
                y=390.0,
                width=155.0,
                height=45.0
            )

        #Network button
            button_image_2 = PhotoImage(
                file=relative_to_assets("button_2_3.png"))
            button_2 = Button(
                image=button_image_2,
                borderwidth=0,
                highlightthickness=0,
                command=lambda: nd.graphnetwork("PERT.xlsx"),
                relief="flat"
            )
            button_2.place(
                x=286.0,
                y=473.0,
                width=155.0,
                height=45.0
            )

        #PERT button
            button_image_3 = PhotoImage(
                file=relative_to_assets("button_3_3.png"))
            button_3 = Button(
                image=button_image_3,
                borderwidth=0,
                highlightthickness=0,
                command=lambda: load_excel(r"PERT.xlsx"),
                relief="flat"
            )
            button_3.place(
                x=286.0,
                y=307.0,
                width=155.0,
                height=45.0
            )

        #Back Button
            button_image_4 = PhotoImage(
                file=relative_to_assets("button_4_3.png"))
            button_4 = Button(
                image=button_image_4,
                borderwidth=0,
                highlightthickness=0,
                command=lambda: screen_2(),
                relief="flat"
            )
            button_4.place(
                x=14.0,
                y=496.0,
                width=90.0,
                height=74.0
            )
            window.resizable(False, False)
            window.mainloop()


        def budgetScreen():
            canvas = Canvas(
                window,
                bg = "#FFFFFF",
                height = 576,
                width = 720,
                bd = 0,
                highlightthickness = 0,
                relief = "ridge"
            )

            canvas.place(x = 0, y = 0)
            image_image_1 = PhotoImage(
                file=relative_to_assets("image_1_4.png"))
            image_1 = canvas.create_image(
                360.0,
                288.0,
                image=image_image_1
            )

            image_image_2 = PhotoImage(
                file=relative_to_assets("image_2_4.png"))
            image_2 = canvas.create_image(
                360.0,
                238.0,
                image=image_image_2
            )

            image_image_3 = PhotoImage(
                file=relative_to_assets("image_3_4.png"))
            image_3 = canvas.create_image(
                450.0,
                65.0,
                image=image_image_3
            )

            image_image_4 = PhotoImage(
                file=relative_to_assets("image_4_4.png"))
            image_4 = canvas.create_image(
                320.0,
                88.0,
                image=image_image_4
            )

#Budget Excel Table Button
            button_image_1 = PhotoImage(
                file=relative_to_assets("button_1_4.png"))
            button_1 = Button(
                image=button_image_1,
                borderwidth=0,
                highlightthickness=0,
                command=lambda: load_excel("Cost_Analysis.xlsx"),
                relief="flat"
            )
            button_1.place(
                x=286.0,
                y=390.0,
                width=155.0,
                height=45.0
            )

#Budget Graphs Button

            button_image_2 = PhotoImage(
                file=relative_to_assets("button_2_4.png"))
            button_2 = Button(
                image=button_image_2,
                borderwidth=0,
                highlightthickness=0,
                command=lambda: Cost.generate_Graphs(),
                relief="flat"
            )
            button_2.place(
                x=286.0,
                y=307.0,
                width=155.0,
                height=45.0
            )

#Back Button
            button_image_3 = PhotoImage(
                file=relative_to_assets("button_3_4.png"))
            button_3 = Button(
                image=button_image_3,
                borderwidth=0,
                highlightthickness=0,
                command=lambda: screen_2(),
                relief="flat"
            )
            button_3.place(
                x=14.0,
                y=496.0,
                width=90.0,
                height=74.0
            )
            #To make the window's dimensions fixed or non-resizable
            window.resizable(False, False)
            #to start the application
            window.mainloop()

            

        window.resizable(False, False)
        window.mainloop()

    window.resizable(False, False)
    window.mainloop()






#To start the main functiomn    
main()

